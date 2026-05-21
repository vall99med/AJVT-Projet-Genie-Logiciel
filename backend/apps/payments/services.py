"""
Logique métier des cotisations, statistiques et export Excel.
"""
import io
import logging
from datetime import datetime

from django.db import transaction
from django.db.models import Sum, Count, Q
from django.utils import timezone
from rest_framework.exceptions import ValidationError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from apps.users.models import User
from apps.members.models import Profile
from apps.dashboard.models import AuditLog
from .models import Payment

logger = logging.getLogger(__name__)


class PaymentService:

    @staticmethod
    def get_member_card(user):
        """Retourne les données de la carte membre digitale."""
        try:
            profile = user.profile
        except Exception:
            profile = None

        current_year = datetime.now().year

        # Détermine le statut de cotisation par ordre de priorité
        cotisation_status = Payment.Status.PENDING
        rejection_reason  = ""
        for status in (Payment.Status.PAID, Payment.Status.SUBMITTED, Payment.Status.REJECTED):
            p = Payment.objects.filter(user=user, year=current_year, status=status).first()
            if p:
                cotisation_status = status
                if status == Payment.Status.REJECTED:
                    rejection_reason = p.rejection_reason
                break

        situation = profile.situation if profile else ""
        if situation == "student":
            detail = " · ".join(filter(None, [
                profile.specialty   if profile else "",
                profile.study_level if profile else "",
            ]))
        elif situation == "employed":
            detail = profile.job_title if profile else ""
        else:
            detail = ""

        return {
            "member_id":        user.id,
            "full_name":        profile.full_name if profile else "",
            "phone":            user.phone,
            "neighborhood":     profile.neighborhood if profile else "",
            "situation":        situation,
            "detail":           detail,
            "cotisation_status": cotisation_status,
            "rejection_reason": rejection_reason,
            "cotisation_year":  current_year,
            "member_since":     user.created_at.year,
        }

    @staticmethod
    def submit_payment(user, validated_data):
        """Crée un paiement SUBMITTED après soumission du reçu par le membre."""
        year = validated_data["year"]

        # Bloque si un paiement non-rejeté existe déjà pour cet user+year
        if Payment.objects.filter(user=user, year=year).exclude(
            status=Payment.Status.REJECTED
        ).exists():
            raise ValidationError("Une demande existe déjà pour cette année.")

        with transaction.atomic():
            payment = Payment.objects.create(
                user            = user,
                year            = year,
                amount          = validated_data["amount"],
                payment_mode    = validated_data["payment_mode"],
                transaction_ref = validated_data.get("transaction_ref", ""),
                receipt_image   = validated_data.get("receipt_image"),
                status          = Payment.Status.SUBMITTED,
                submitted_at    = timezone.now(),
            )
            AuditLog.log(
                "payment_submitted",
                performed_by=user,
                target_user=user,
                details={
                    "year":   year,
                    "amount": str(validated_data["amount"]),
                    "mode":   validated_data["payment_mode"],
                },
            )
        payment.compress_receipt()
        logger.info("Reçu soumis par %s — année %s", user.phone, year)
        return payment

    @staticmethod
    def review_payment(payment_id, action, rejection_reason, reviewed_by):
        """Approuve ou rejette un reçu soumis par un membre."""
        try:
            payment = Payment.objects.select_related("user").get(pk=payment_id)
        except Payment.DoesNotExist:
            raise ValidationError("Paiement introuvable.")

        if payment.status != Payment.Status.SUBMITTED:
            raise ValidationError("Ce paiement n'est pas en attente de vérification.")

        now = timezone.now()
        with transaction.atomic():
            if action == "approve":
                payment.status      = Payment.Status.PAID
                payment.paid_at     = now
                payment.reviewed_by = reviewed_by
                payment.reviewed_at = now
                payment.save()
                AuditLog.log(
                    "payment_approved",
                    performed_by=reviewed_by,
                    target_user=payment.user,
                    details={"payment_id": payment_id, "year": payment.year},
                )
                logger.info("Paiement %s approuvé par %s", payment_id, reviewed_by.phone)
            else:
                payment.status           = Payment.Status.REJECTED
                payment.rejection_reason = rejection_reason
                payment.reviewed_by      = reviewed_by
                payment.reviewed_at      = now
                payment.save()
                AuditLog.log(
                    "payment_rejected",
                    performed_by=reviewed_by,
                    target_user=payment.user,
                    details={"payment_id": payment_id, "year": payment.year, "reason": rejection_reason},
                )
                logger.info("Paiement %s rejeté par %s", payment_id, reviewed_by.phone)

        return payment

    @staticmethod
    def get_submitted_payments():
        """Retourne tous les reçus soumis en attente de vérification (admin)."""
        return (
            Payment.objects
            .filter(status=Payment.Status.SUBMITTED)
            .select_related("user", "user__profile")
            .order_by("submitted_at")
        )

    @staticmethod
    def create_payment(user_id, validated_data, created_by):
        """Enregistre directement une cotisation cash (admin). Bloque les doublons actifs."""
        try:
            user = User.objects.get(pk=user_id)
        except User.DoesNotExist:
            raise ValidationError("Membre introuvable.")

        if user.status != User.Status.ACTIVE:
            raise ValidationError("Ce compte n'est pas actif.")

        year = validated_data["year"]
        # Bloque si un paiement payé ou soumis existe déjà
        if Payment.objects.filter(
            user=user, year=year, status__in=[Payment.Status.PAID, Payment.Status.SUBMITTED]
        ).exists():
            raise ValidationError("Cotisation déjà enregistrée ou en cours de traitement pour cette année.")

        with transaction.atomic():
            payment = Payment.objects.create(
                user         = user,
                year         = year,
                amount       = validated_data["amount"],
                payment_mode = validated_data.get("payment_mode", Payment.PaymentMode.CASH),
                notes        = validated_data.get("notes", ""),
                status       = Payment.Status.PAID,
                paid_at      = timezone.now(),
                validated_by = created_by,
            )
            AuditLog.log(
                "payment_created",
                performed_by=created_by,
                target_user=user,
                details={"year": year, "amount": str(validated_data["amount"])},
            )
        logger.info("Cotisation créée pour %s — année %s par %s", user.phone, year, created_by.phone)
        return payment

    @staticmethod
    def get_member_payments(user):
        """Retourne l'historique des cotisations d'un membre, du plus récent au plus ancien."""
        return Payment.objects.filter(user=user).order_by("-year")

    @staticmethod
    def get_all_payments(filters=None):
        """Retourne toutes les cotisations avec filtres optionnels : year, status."""
        qs = Payment.objects.select_related("user", "user__profile").order_by("-created_at")
        if filters:
            if filters.get("year"):
                qs = qs.filter(year=filters["year"])
            if filters.get("status"):
                qs = qs.filter(status=filters["status"])
        return qs

    @staticmethod
    def get_dashboard_stats():
        """Calcule les statistiques du tableau de bord via agrégations Django (sans boucles)."""
        current_year = datetime.now().year

        user_counts = User.objects.aggregate(
            total    = Count("id", filter=Q(status=User.Status.ACTIVE)),
            pending  = Count("id", filter=Q(status=User.Status.PENDING)),
            rejected = Count("id", filter=Q(status=User.Status.REJECTED)),
        )

        payment_stats = Payment.objects.filter(
            year=current_year, status=Payment.Status.PAID
        ).aggregate(
            total_amount  = Sum("amount"),
            members_paid  = Count("user", distinct=True),
        )

        total_members = user_counts["total"] or 0
        members_paid  = payment_stats["members_paid"] or 0

        by_situation = dict(
            Profile.objects.filter(user__status=User.Status.ACTIVE)
            .values("situation")
            .annotate(count=Count("id"))
            .values_list("situation", "count")
        )

        return {
            "total_members":           total_members,
            "pending_members":         user_counts["pending"] or 0,
            "rejected_members":        user_counts["rejected"] or 0,
            "total_cotisations_year":  str(payment_stats["total_amount"] or 0),
            "members_paid_year":       members_paid,
            "members_pending_payment": total_members - members_paid,
            "by_situation": {
                "student":    by_situation.get("student",    0),
                "employed":   by_situation.get("employed",   0),
                "unemployed": by_situation.get("unemployed", 0),
            },
        }

    @staticmethod
    def export_members_excel():
        """Génère un fichier Excel des membres actifs et retourne ses bytes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Membres AJVT"

        headers     = ["Nom complet", "Téléphone", "Situation", "Spécialité/Poste",
                       "Région", "Statut cotisation", "Année adhésion"]
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

        for col, header in enumerate(headers, start=1):
            cell       = ws.cell(row=1, column=col, value=header)
            cell.font  = header_font
            cell.fill  = header_fill

        current_year = datetime.now().year
        profiles = (
            Profile.objects.filter(user__status=User.Status.ACTIVE)
            .select_related("user")
            .order_by("full_name")
        )
        paid_users = set(
            Payment.objects.filter(year=current_year, status=Payment.Status.PAID)
            .values_list("user_id", flat=True)
        )

        for profile in profiles:
            specialty  = profile.specialty or profile.job_title
            cotisation = "Payée" if profile.user_id in paid_users else "En attente"
            ws.append([
                profile.full_name,
                profile.user.phone,
                profile.get_situation_display(),
                specialty,
                profile.neighborhood,
                cotisation,
                profile.user.created_at.year,
            ])

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
